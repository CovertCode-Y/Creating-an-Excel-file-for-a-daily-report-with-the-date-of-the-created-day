import os
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime

def create_daily_log(ws, date):
    ws.title = date.strftime("%Y-%m-%d")
    ws.cell(row=1, column=1, value=f"Daily Task Report - {date.strftime('%d %B %Y')}").font = Font(size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    headers = ["Time", "Task", "Description", "Duration (Hours)", "Notes"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col, value=header).font = Font(bold=True)

    for row in range(3, 28):  # Assuming 25 rows for tasks
        for col in range(1, 6):
            ws.cell(row=row, column=col).alignment = Alignment(vertical='top')

def create_log_file():
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    folder_name = "Daily Task Report"
    folder_path = os.path.join(desktop, folder_name)
    os.makedirs(folder_path, exist_ok=True)

    wb = Workbook()
    current_date = datetime.now()
    create_daily_log(wb.active, current_date)

    file_path = os.path.join(folder_path, f"Daily_Task_Report_{current_date.strftime('%Y%m%d')}.xlsx")
    wb.save(file_path)
    print(f"Folder '{folder_name}' created on Desktop and Excel file for {current_date.strftime('%d %B %Y')} saved inside.")

    # Open the folder containing the file
    if os.name == 'nt':  # For Windows
        os.startfile(folder_path)
    elif os.name == 'posix':  # For macOS and Linux
        subprocess.call(['open', folder_path])

if __name__ == "__main__":
    create_log_file()
